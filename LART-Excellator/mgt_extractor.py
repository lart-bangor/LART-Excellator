# importing packages
import pandas as pd
import json
import os,glob
import openpyxl
from openpyxl.styles import Font
from rich.prompt import Prompt 
import time
import navigate


#REVERSAL
#for the purposes of statistical analysis, some adective pairs indicate a positive response when rated high, while others do so when rated low
# values to be reversed are:
reversal = ["amusing", "ignorant", "pretentious"]


#lists of guises for which ratings are to be exracted.
majList = ['s1_maj_ratings', 's2_maj_ratings','s3_maj_ratings','s4_maj_ratings']
rmlList = ['s1_rml_ratings', 's2_rml_ratings','s3_rml_ratings','s4_rml_ratings']

def label_sheet(wb, sheetName, old):
    sheet = wb[old]
    sheet.title = sheetName
    wb.active = wb[sheetName]

    c1 = sheet.cell(row = 1, column = 1)
    c2 = sheet.cell(row = 1, column = 2)
    c3 = sheet.cell(row = 1, column = 3)
    c4 = sheet.cell(row = 1, column = 4)
    c5 = sheet.cell(row = 1, column = 5)
    c6 = sheet.cell(row = 1, column = 6)
    c7 = sheet.cell(row = 1, column = 7)
    c8 = sheet.cell(row = 1, column = 8)
    c9 = sheet.cell(row = 1, column = 9)
    c10 = sheet.cell(row = 1, column = 10)
    c11 = sheet.cell(row = 1, column = 11)
    c12 = sheet.cell(row = 1, column = 12)
    c13 = sheet.cell(row = 1, column = 13)
    c14 = sheet.cell(row = 1, column = 14)
    c15 = sheet.cell(row = 1, column = 15)
    c16 = sheet.cell(row = 1, column = 16)
    c17 = sheet.cell(row = 1, column = 17)
    c18 = sheet.cell(row = 1, column = 18)
    c19 = sheet.cell(row = 1, column = 19)
    c20 = sheet.cell(row = 1, column = 20)

    c1.value = "Participant ID"
    sheet.column_dimensions['A'].width = 14
    c1.font = Font(bold=True)
    c2.value = "Stimulus voice"
    sheet.column_dimensions['B'].width = 14
    c2.font = Font(bold=True)
    c3.value = "amusing (rev)"
    c4.value = "open-minded"
    c5.value = "attractive"
    c6.value = "trustworthy"
    c7.value = "ignorant (rev)"
    c8.value = "polite"
    c9.value = "ambitious"
    c10.value = "international"
    c11.value = "cool"
    c12.value = "intelligent"
    c13.value = "influential"
    c14.value = "liekeable"
    c15.value = "educated"
    c16.value = "friendly"
    c17.value = "honest"
    c18.value = "competent"
    c19.value = "natural"
    c20.value = "pretentious (rev)"
    return sheet

def reversable(adjective, number):
    """Reverses value of ajectives that appear in list called 'reversal'"""
    value = float(number)
    if adjective in reversal:
       # print(f'{adjective} is present in the list')
       # print("Value was", value)
        final_value = 100-value
       # print("Value IS", final_value)
    else:
        final_value = value
       # print(f'{adjective} No No No in List')
    return final_value

def main(folder_path, thisPath, location):
    langLabels = location.split("_")
    rmlLabel = langLabels[0]
    majLabel = langLabels[1]

    wb = openpyxl.Workbook()
    wb.create_sheet("sheet1")
    wb.create_sheet("sheet2")
         
    columnNumData = 3
    rowNumData = 2          #start at 2 because 1 is labels
    rowNumDataRml = 2          #start at 2 because 1 is labels

    sheetMaj = label_sheet(wb, 'majLang' + '_' + majLabel, 'sheet1')
    sheetRml = label_sheet(wb, 'rml' + '_' + rmlLabel, 'sheet2')

    for filename in glob.glob(os.path.join(folder_path, '*.json')):
                  
        with open(filename, 'r') as f:
            data = f.read()
            if data == "":
                print("\nNO CONTENT\n")
            navigate.processing_info(filename)
            json_object = json.loads(data)
            meta = json_object['meta']
            print(f'META IS: {meta}\n')
            p_id = meta['participant_id']
            
            for item in majList:
                print(f'MAJ Item is: {item}')
                maj = json_object[item]
                print(f'ROW NUMBER is: {str(rowNumData)}\n')
                print(f'json object is: {maj}')
                
                for key in maj:
                    cell_f = sheetMaj.cell(row = rowNumData, column = 1)
                    cell_f.value = p_id
                    cell_stim = sheetMaj.cell(row = rowNumData, column = 2)
                    cell_stim.value = item
                    cell = sheetMaj.cell(row = rowNumData, column = columnNumData)
                    final_value = reversable(key, maj[key])
                    #cell.value = float(maj[key])
                    cell.value = final_value
                    columnNumData +=1
                rowNumData +=1
                columnNumData = 3
              
            for item in rmlList:
                print(f'RML Item is: {item}')
                rml = json_object[item]
                print(f'ROW NUMBER is: {str(rowNumDataRml)}\n')
                print(f'json object is: {rml}')
                for key in rml:
                    cell_f = sheetRml.cell(row = rowNumDataRml, column = 1)
                    cell_f.value = p_id
                    cell_stim = sheetRml.cell(row = rowNumDataRml, column = 2)
                    cell_stim.value = item
                    cell = sheetRml.cell(row = rowNumDataRml, column = columnNumData)
                    final_value = reversable(key, rml[key])
                    cell.value = final_value
                    #cell.value = float(rml[key])
                    columnNumData +=1
                rowNumDataRml+=1
                columnNumData = 3
    output_file = "mgtData" + location + ".xlsx"
    output_path = os.path.join(thisPath, "outputs/")
    print(f' this path is {thisPath}')
    output_fullPath = os.path.join(output_path, output_file)
    wb.save(output_fullPath)
    navigate.completed_info("MGT", output_file)
    
def exract_mgt(path, directoryPath):
    location = Prompt.ask('Enter the code for the data to be extracted. Default is: CYM_ENG',
                        default='CYM_ENG',
                        choices=['LMO_IT', 'LtzGer_GerBE', 'exit'])
    if location == "exit":
            navigate.abort()
    navigate.data_selection(location)
    
    folder_path = os.path.join(path, location)
    if os.path.isdir(folder_path):
        if  os.listdir(folder_path):
           # navigate.selection(folder_path)
            time.sleep(1.0)
            main(folder_path, directoryPath, location)
            print("\n")
        else:
            navigate.is_empty(folder_path)            
    else:
        navigate.locate(directoryPath, folder_path)
        navigate.not_found(location)
