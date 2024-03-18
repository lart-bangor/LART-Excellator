
# importing packages
import time
import pandas as pd
import json
import os,glob
import openpyxl
from rich.prompt import Prompt 
import sys

#REVERSAL
#for the purposes of statistical analysis, some adective pairs indicate a positive response when rated high, while others do so when rated low
# below is the full set with each value

# adjectives": {									                     *POSITIVE*
#         "logic":        ["logical",         "illogical"],					low
#         "elegance":     ["inelegant",       "elegant"],					high
#         "fluency":      ["choppy",          "fluent"],					high
#         "ambiguity":    ["unambiguous",     "ambiguous"],					low
#         "appeal":       ["appealing",       "abhorrent"],					low
#         "structure":    ["unstructured",    "structured"],				high
#         "precision":    ["precise",         "vague"],						low
#         "harshness":    ["harsh",           "soft"],						high
#         "flow":         ["flowing",         "abrupt"],					low
#         "beauty":       ["beautiful",       "ugly"],						low
#         "sistem":       ["systematic",      "unsystematic"],				low
#         "pleasure":     ["pleasant",        "unpleasant"],				low
#         "smoothness":   ["smooth",          "raspy"],						low
#         "grace":        ["clumsy",          "graceful"],					high
#         "angularity":   ["angular",         "round"]						high


#as above, these are the adjectives whose value must be reveresed, called by reversable()
reversal = ["appeal", "beauty", "pleasure", "flow", "smoothness", "logic", "precision", "sistem", "ambiguity"]

def label_sheet(wb, sheetName, old):
    sheet = wb[old]
    sheet.title = sheetName
    wb.active = wb[sheetName]

    c1 = sheet.cell(row = 1, column = 1)
    c2 = sheet.cell(row = 1, column = 2)
    c3 = sheet.cell(row = 1, column = 3)
    c4 = sheet.cell(row = 1, column = 4)
    c5 = sheet.cell(row = 1, column = 5)
    c6 = sheet.cell(row = 1, column = 6)
    c7 = sheet.cell(row = 1, column = 7)
    c8 = sheet.cell(row = 1, column = 8)
    c9 = sheet.cell(row = 1, column = 9)
    c10 = sheet.cell(row = 1, column = 10)
    c11 = sheet.cell(row = 1, column = 11)
    c12 = sheet.cell(row = 1, column = 12)
    c13 = sheet.cell(row = 1, column = 13)
    c14 = sheet.cell(row = 1, column = 14)
    c15 = sheet.cell(row = 1, column = 15)
    c16 = sheet.cell(row = 1, column = 16)

    c1.value = "Filename"
    c2.value = "ambiguity (rev)"
    c3.value = "angularity"
    c4.value = "appeal (rev)"
    c5.value = "beauty (rev)"
    c6.value = "elegance"
    c7.value = "flow (rev)"
    c8.value = "fluency"
    c9.value = "grace"
    c10.value = "harshness"
    c11.value = "logic (rev)"
    c12.value = "pleasure (rev)"
    c13.value = "precision (rev)"
    c14.value = "system (rev)"
    c15.value = "smoothness (rev)"
    c16.value = "structure"
    return sheet

def reversable(header, number):
    """Reverses value of ajectives that appear in list called 'reversal'"""
    value = float(number)
    splitHeaer = header.split("_")
    adjective = splitHeaer[2]
    if adjective in reversal:
        #print(f'{adjective} is present in the list')
        #print("Value was", value)
        final_value = 100-value
        #print("Value IS", value)
    else:
        final_value = value
        #print(f'{adjective} No No No in List')
    return final_value

def abort():
    print("\nOperation aborted")
    sys.exit()


def main(folder_path, thisPath, location):
    langLabels = location.split("_")
    rmlLabel = langLabels[0]
    majLabel = langLabels[1]

    wb = openpyxl.Workbook()
    wb.create_sheet("sheet1")
    wb.create_sheet("sheet2")
         
    columnNumData = 2
    rowNumData = 2          #start at 2 because 1 is labels
    sheetMaj = label_sheet(wb, 'majLang' + '_' + majLabel, 'sheet1')
    sheetRml = label_sheet(wb, 'rml' + '_' + rmlLabel, 'sheet2')

    for filename in glob.glob(os.path.join(folder_path, '*.json')):
      # with io.open(filename, encoding="utf8") as f:
       # with open(filename, 'r') as f:
        with open(filename, encoding="latin-1") as f:
            text = f.read()
            print(f'\n Currently processing: {filename}\n')
            json_object = json.loads(text)
            maj = json_object["Ratings_atolRatingsMaj"]
            rml = json_object["Ratings_atolRatingsRml"]
            
            for key in maj:
                cell_f = sheetMaj.cell(row = rowNumData, column = 1)
                cell_f.value = filename
                cell = sheetMaj.cell(row = rowNumData, column = columnNumData)
                final_value = reversable(key, maj[key])
                cell.value = final_value
                columnNumData +=1
            columnNumData = 2
            
            for key in rml:
                cell_f = sheetRml.cell(row = rowNumData, column = 1)
                cell_f.value = filename
                cell = sheetRml.cell(row = rowNumData, column = columnNumData)
                final_value = reversable(key, rml[key])
                cell.value = final_value
                columnNumData +=1
            rowNumData+=1
            columnNumData = 2
    wb.save(thisPath + "\\atolData" + location + ".xlsx")

def extract_atol(path, directoryPath):

    location = Prompt.ask('Enter the code for the data to be extracted. Default is: CYM_ENG',
                        default='CYM_ENG',
                        choices=['LMO_IT', 'LtzGer_GerBE', 'exit'])
    if location == "exit":
            abort()
    print(f'\n\t\t\tData folder selected: {location}')
    folder_path = os.path.join(path, location)
    if os.path.isdir(folder_path):
        if  os.listdir(folder_path):
            print(f'\n You have selected to work in the following folder: {folder_path}\n')
            time.sleep(2.5)
            main(folder_path, directoryPath, location)
            print("\n")
        else:
            print(f'\n\tEMPTY DIRECTORY: The folder \ {folder_path} \ is empty.\n')
    else:
           print(f'\n\tNOT FOUND: There is no folder named {location} inside {path}\n\t Please check folder and try again.\n')
